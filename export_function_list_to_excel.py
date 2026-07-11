# -*- coding: utf-8 -*-
"""将 D05-系统功能清单.md 导出为 Excel。"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
MD_PATH = ROOT / "D05-系统功能清单.md"
OUT_PATH = ROOT / "系统功能清单_V2.4终版.xlsx"

MODULE_COLS = [
    "编号",
    "逻辑域",
    "功能模块",
    "功能描述",
    "主要功能点",
    "交付级别",
    "验收要点",
]

MODULE_HEADERS = [
    "一级章节",
    "二级章节",
    "三级章节",
    "四级章节",
    *MODULE_COLS,
]


def clean_cell(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    return text


def parse_md(path: Path):
    lines = path.read_text(encoding="utf-8").splitlines()
    modules = []
    h1, h2, h3, h4 = "", "", "", ""

    for line in lines:
        if line.startswith("## "):
            h1 = clean_cell(line[3:])
            h2, h3, h4 = "", "", ""
            continue
        if line.startswith("### "):
            h2 = clean_cell(line[4:])
            h3, h4 = "", ""
            continue
        if line.startswith("#### "):
            h3 = clean_cell(line[5:])
            h4 = ""
            continue

        if not line.startswith("| M"):
            continue
        if re.match(r"^\|\s*[-:]+", line):
            continue

        parts = [p.strip() for p in line.strip().strip("|").split("|")]
        if len(parts) < 7 or not parts[0].startswith("M"):
            continue

        modules.append(
            {
                "一级章节": h1,
                "二级章节": h2,
                "三级章节": h3,
                "四级章节": h4,
                **dict(zip(MODULE_COLS, [clean_cell(p) for p in parts[:7]])),
            }
        )

    return lines, modules


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, max_width))


def write_sheet(ws, headers, rows, wrap_cols=None):
    wrap_cols = wrap_cols or set()
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def main():
    lines, modules = parse_md(MD_PATH)
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "文档信息"
    meta = [
        ("文档名称", "承德高新区智慧城市基础平台 — 系统功能清单"),
        ("文档版本", "V2.4（终版）"),
        ("编制日期", "2026-06-22"),
        ("模块总数", str(len(modules))),
        ("来源文件", MD_PATH.name),
    ]
    for i, (k, v) in enumerate(meta, 1):
        ws_info.cell(row=i, column=1, value=k)
        ws_info.cell(row=i, column=2, value=v)
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 50

    ws_mod = wb.create_sheet("功能模块清单")
    mod_rows = [[m[h] for h in MODULE_HEADERS] for m in modules]
    write_sheet(ws_mod, MODULE_HEADERS, mod_rows, wrap_cols={5, 7, 8, 10, 11})

    summary_headers = [
        "功能清单章节",
        "逻辑域",
        "V3.0原文完整板块名称",
        "内部简称",
        "M范围",
        "模块数",
    ]
    summary_rows = []
    for line in lines:
        if (
            line.startswith("| 一、")
            or line.startswith("| 二、")
            or line.startswith("| 三、")
            or line.startswith("| 四、")
            or line.startswith("| 五、")
            or line.startswith("| **合计**")
        ):
            cells = [clean_cell(c) for c in line.strip().strip("|").split("|")]
            if len(cells) >= 6:
                summary_rows.append(cells[:6])

    ws_sum = wb.create_sheet("汇总统计")
    write_sheet(ws_sum, summary_headers, summary_rows)

    level_rows = []
    for line in lines:
        if line.startswith("| **L1**") or line.startswith("| **L2**") or line.startswith(
            "| **L3**"
        ):
            cells = [clean_cell(c) for c in line.strip().strip("|").split("|")]
            if len(cells) >= 3:
                level_rows.append(cells[:3])
    if level_rows:
        ws_level = wb.create_sheet("交付级别分布")
        write_sheet(ws_level, ["级别", "数量", "说明"], level_rows, wrap_cols={3})

    ms_rows = []
    for line in lines:
        if line.startswith("| **MS"):
            cells = [clean_cell(c) for c in line.strip().strip("|").split("|")]
            if len(cells) >= 3:
                ms_rows.append(cells[:3])
    ws_ms = wb.create_sheet("交付里程碑")
    write_sheet(ws_ms, ["里程碑", "内容", "关联M范围"], ms_rows, wrap_cols={2, 3})

    pending_rows = []
    for line in lines:
        if re.match(r"^\|\s*\d+\s*\|", line):
            cells = [clean_cell(c) for c in line.strip().strip("|").split("|")]
            if len(cells) >= 4:
                pending_rows.append(cells[:4])
    ws_pending = wb.create_sheet("待确认事项")
    write_sheet(
        ws_pending,
        ["序号", "待确认项", "影响模块示例", "备注"],
        pending_rows,
        wrap_cols={2, 3, 4},
    )

    wb.save(OUT_PATH)
    print(f"已导出: {OUT_PATH}")
    print(f"模块行数: {len(modules)}")


if __name__ == "__main__":
    main()
