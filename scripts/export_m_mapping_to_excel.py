# -*- coding: utf-8 -*-
"""将 Mxxx 实现映射表 Markdown 导出为 Excel。"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "D06-Mxxx实现映射表.md"
OUT_PATH = ROOT / "Mxxx实现映射表_V1.0.xlsx"

MAPPING_HEADERS = [
    "M",
    "功能模块",
    "逻辑域",
    "L",
    "V3板块",
    "实现方式",
    "框架/组件",
    "部署",
    "门户集成",
    "代码包",
    "二次开发要点",
]


def clean_cell(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    return text


def parse_mapping_table(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    in_section = False
    rows: list[list[str]] = []
    for line in lines:
        if line.startswith("## 三、M001"):
            in_section = True
            continue
        if in_section and line.startswith("## "):
            break
        if not in_section:
            continue
        if not re.match(r"^\| M\d{3} \|", line):
            continue
        if re.match(r"^\|\s*[-:]+", line):
            continue
        parts = [clean_cell(p) for p in line.strip().strip("|").split("|")]
        if len(parts) >= 11:
            rows.append(parts[:11])
    rows.sort(key=lambda r: int(r[0][1:]))
    return rows


def parse_stats_table(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    in_section = False
    rows: list[list[str]] = []
    for line in lines:
        if line.startswith("### 2.1"):
            in_section = True
            continue
        if in_section and line.startswith("###"):
            break
        if not in_section:
            continue
        if not line.startswith("|") or "实现方式" in line or re.match(r"^\|\s*[-:]+", line):
            continue
        parts = [clean_cell(p) for p in line.strip().strip("|").split("|")]
        if len(parts) >= 2:
            rows.append(parts[:2])
    return rows


def parse_framework_index(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    in_section = False
    rows: list[list[str]] = []
    for line in lines:
        if line.startswith("## 四、框架"):
            in_section = True
            continue
        if in_section and line.startswith("## "):
            break
        if not in_section:
            continue
        if not line.startswith("|") or "框架/组件" in line or re.match(r"^\|\s*[-:]+", line):
            continue
        parts = [clean_cell(p) for p in line.strip().strip("|").split("|")]
        if len(parts) >= 3:
            rows.append(parts[:3])
    return rows


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, max_width))


def write_sheet(ws, headers: list[str], rows: list[list[str]], wrap_cols: set[int] | None = None) -> None:
    wrap_cols = wrap_cols or set()
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def main() -> int:
    if not MD_PATH.exists():
        print(f"未找到 {MD_PATH}，请先运行 scripts/gen_m_mapping.py")
        return 1

    mapping_rows = parse_mapping_table(MD_PATH)
    if len(mapping_rows) != 215:
        print(f"映射表行数异常: {len(mapping_rows)}，期望 215")
        return 1

    stats_rows = parse_stats_table(MD_PATH)
    index_rows = parse_framework_index(MD_PATH)

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "文档信息"
    meta = [
        ("文档名称", "Mxxx 实现映射表"),
        ("文档版本", "V1.0（终稿）"),
        ("模块总数", str(len(mapping_rows))),
        ("配套清单", "系统功能清单 V2.6"),
        ("来源文件", MD_PATH.name),
        ("输出文件", OUT_PATH.name),
    ]
    for i, (k, v) in enumerate(meta, 1):
        ws_info.cell(row=i, column=1, value=k)
        ws_info.cell(row=i, column=2, value=v)
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 50

    ws_map = wb.create_sheet("M001-M215映射")
    write_sheet(ws_map, MAPPING_HEADERS, mapping_rows, wrap_cols={5, 6, 10, 11})

    ws_stats = wb.create_sheet("实现方式统计")
    write_sheet(ws_stats, ["实现方式", "数量"], stats_rows)

    ws_idx = wb.create_sheet("框架索引")
    write_sheet(ws_idx, ["框架/组件", "模块数", "M 编号"], index_rows, wrap_cols={3})

    wb.save(OUT_PATH)
    print(f"已导出: {OUT_PATH}")
    print(f"映射行数: {len(mapping_rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
