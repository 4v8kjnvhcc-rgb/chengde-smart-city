# -*- coding: utf-8 -*-
"""Import 公安业务 CSV into source-mysql gongan_biz schema."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pandas as pd
import pymysql


def sql_type(dtype, length) -> str:
    t = (dtype or "string").lower().strip()
    try:
        n = int(str(length).split(".")[0]) if length not in (None, "", "None") else 64
    except Exception:
        n = 64
    if t in ("int", "integer"):
        return "INT"
    if t == "date":
        return "DATE"
    if t in ("datetime", "timestamp"):
        return "DATETIME"
    if t in ("decimal", "number", "float", "double"):
        return "DECIMAL(18, 2)"
    n = max(n, 8)
    if n > 512:
        n = 512
    return f"VARCHAR({max(n, 32)})"


def main() -> None:
    desktop = Path(os.environ["USERPROFILE"]) / "Desktop"
    base = Path(r"C:\Users\ming\WorkBuddy\2026-07-19-10-19-49")

    meta_map: dict[str, Path] = {}
    csv_map: dict[str, Path] = {}
    for xf in desktop.glob("*.xlsx"):
        if "常住人口" in xf.name:
            meta_map["resident"] = xf
        elif "治安重点" in xf.name:
            meta_map["unit"] = xf
    for p in base.rglob("*.csv"):
        if "常住人口" in p.name:
            csv_map["resident"] = p
        elif "治安重点" in p.name:
            csv_map["unit"] = p

    tables = {
        "resident": {
            "table": "ga_rk_resident",
            "comment": "公安分局_常住人口登记表",
            "meta": meta_map["resident"],
            "csv": csv_map["resident"],
        },
        "unit": {
            "table": "ga_za_key_unit",
            "comment": "公安分局_治安重点单位表",
            "meta": meta_map["unit"],
            "csv": csv_map["unit"],
        },
    }

    conn = pymysql.connect(
        host="127.0.0.1",
        port=3308,
        user="root",
        password="source_root",
        charset="utf8mb4",
        autocommit=True,
    )
    cur = conn.cursor()
    cur.execute("CREATE DATABASE IF NOT EXISTS gongan_biz CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
    cur.execute("CREATE USER IF NOT EXISTS 'gongan'@'%' IDENTIFIED BY 'Gongan@2026'")
    cur.execute("GRANT SELECT, INSERT, UPDATE, DELETE ON gongan_biz.* TO 'gongan'@'%'")
    cur.execute("CREATE USER IF NOT EXISTS 'probe'@'%' IDENTIFIED BY 'probe_pass'")
    cur.execute("GRANT SELECT ON gongan_biz.* TO 'probe'@'%'")
    cur.execute("FLUSH PRIVILEGES")
    cur.execute("USE gongan_biz")

    results = {}
    for cfg in tables.values():
        wb = openpyxl.load_workbook(cfg["meta"], data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        cols = []
        for r in rows[1:]:
            if not r or not r[1]:
                continue
            en, zh, dtype, length, pk = r[1], r[2], r[4], r[5], r[8]
            cols.append((en, zh, dtype, length, str(pk).upper() == "Y"))

        pk = "record_id" if any(c[0] == "record_id" for c in cols) else cols[0][0]
        col_defs = []
        for en, zh, dtype, length, _ in cols:
            st = sql_type(dtype, length)
            comment = (zh or en).replace("'", "")
            nn = "NOT NULL" if en == pk else "NULL"
            col_defs.append(f"  `{en}` {st} {nn} COMMENT '{comment}'")

        cur.execute(f"DROP TABLE IF EXISTS `{cfg['table']}`")
        create_sql = (
            f"CREATE TABLE `{cfg['table']}` (\n"
            + ",\n".join(col_defs)
            + f",\n  PRIMARY KEY (`{pk}`)\n"
            + f") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='{cfg['comment']}'"
        )
        cur.execute(create_sql)

        enc = "utf-8"
        blob = cfg["csv"].read_bytes()[:4000]
        for e in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                blob.decode(e)
                enc = e
                break
            except Exception:
                pass

        df = pd.read_csv(cfg["csv"], encoding=enc, dtype=str)
        keep = [c[0] for c in cols if c[0] in df.columns]
        df = df[keep]
        placeholders = ",".join(["%s"] * len(keep))
        colnames = ",".join(f"`{c}`" for c in keep)
        insert_sql = f"INSERT INTO `{cfg['table']}` ({colnames}) VALUES ({placeholders})"
        data = []
        for row in df.itertuples(index=False, name=None):
            cleaned = []
            for v in row:
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    cleaned.append(None)
                else:
                    s = str(v).strip()
                    cleaned.append(None if s == "" or s.lower() == "nan" else s)
            data.append(tuple(cleaned))
        for i in range(0, len(data), 100):
            cur.executemany(insert_sql, data[i : i + 100])

        cur.execute(f"SELECT COUNT(*) FROM `{cfg['table']}`")
        cnt = int(cur.fetchone()[0])
        results[cfg["table"]] = {
            "rows": cnt,
            "columns": len(keep),
            "comment": cfg["comment"],
            "pk": pk,
        }
        print(f"OK {cfg['table']} rows={cnt} cols={len(keep)}")

    cur.execute("SHOW TABLES")
    print("TABLES", [r[0] for r in cur.fetchall()])
    cur.close()
    conn.close()
    print("RESULT", results)


if __name__ == "__main__":
    main()
