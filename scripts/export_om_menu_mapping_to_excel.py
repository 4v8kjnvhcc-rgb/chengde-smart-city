# -*- coding: utf-8 -*-
"""将 D10-OM菜单名映射表.md 导出为 Excel。"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "D10-OM菜单名映射表.md"
OUT_PATH = ROOT / "D10-OM菜单名映射表_V1.0.xlsx"

HEADERS = [
    "M",
    "功能模块",
    "门户父菜单",
    "门户菜单名",
    "门户路由",
    "OM 原生入口",
    "OM 路由/API",
    "集成方式",
    "MS",
    "备注",
]


def clean_cell(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    return text


def parse_mapping_rows(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    in_section = False
    rows: list[list[str]] = []
    for line in lines:
        if line.startswith("## 四、逐条映射表"):
            in_section = True
            continue
        if in_section and line.startswith("## 五、"):
            break
        if not in_section:
            continue
        if not re.match(r"^\| M\d{3} \|", line):
            continue
        if re.match(r"^\|\s*[-:]+", line):
            continue
        parts = [clean_cell(p) for p in line.strip().strip("|").split("|")]
        if len(parts) >= 10:
            rows.append(parts[:10])
    rows.sort(key=lambda r: int(r[0][1:]))
    return rows


def parse_neighbor_rows(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    in_section = False
    rows: list[list[str]] = []
    for line in lines:
        if line.startswith("## 六、非 OM"):
            in_section = True
            continue
        if in_section and line.startswith("## 七、"):
            break
        if not in_section:
            continue
        if not re.match(r"^\| M\d{3} \|", line):
            continue
        if re.match(r"^\|\s*[-:]+", line):
            continue
        parts = [clean_cell(p) for p in line.strip().strip("|").split("|")]
        if len(parts) >= 4:
            rows.append(parts[:4])
    return rows


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def main() -> int:
    if not MD_PATH.exists():
        print(f"缺少 {MD_PATH}")
        return 1

    mapping = parse_mapping_rows(MD_PATH)
    neighbors = parse_neighbor_rows(MD_PATH)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "OM菜单映射"
    ws1.append(HEADERS)
    style_header(ws1)
    for row in mapping:
        ws1.append(row)
    auto_width(ws1)

    ws2 = wb.create_sheet("非OM邻域索引")
    ws2.append(["M", "功能模块", "实现方式", "门户归属"])
    style_header(ws2)
    for row in neighbors:
        ws2.append(row)
    auto_width(ws2)

    wb.save(OUT_PATH)
    print(f"已导出 {OUT_PATH}（OM {len(mapping)} 条，邻域 {len(neighbors)} 条）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
