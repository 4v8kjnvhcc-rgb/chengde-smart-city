#!/usr/bin/env python3
"""Normalize 功能清单.xlsx (forward-fill merged cells) and emit scope-specific JSON."""
from __future__ import annotations

import argparse
import json
import os
import re
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl required: pip install openpyxl") from exc

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "requirements", "功能清单.xlsx")
CATALOG_DIR = os.path.join(ROOT, "catalog")

COL_PLATFORM = 2
COL_SYSTEM = 3
COL_SUBSYSTEM = 4
COL_MODULE = 5
COL_DESC = 6
HEADER_ROW = 2
DATA_START = 3


def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def load_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    carry = {"platform": "", "system": "", "subsystem": "", "moduleName": ""}
    for r in range(DATA_START, ws.max_row + 1):
        platform = _s(ws.cell(r, COL_PLATFORM).value) or carry["platform"]
        system = _s(ws.cell(r, COL_SYSTEM).value) or carry["system"]
        subsystem = _s(ws.cell(r, COL_SUBSYSTEM).value) or carry["subsystem"]
        module_name = _s(ws.cell(r, COL_MODULE).value) or carry["moduleName"]
        desc = _s(ws.cell(r, COL_DESC).value)
        if not any([platform, system, subsystem, module_name, desc]):
            continue
        carry = {
            "platform": platform,
            "system": system,
            "subsystem": subsystem,
            "moduleName": module_name,
        }
        rows.append(
            {
                "rowIndex": r,
                "platform": platform,
                "system": system,
                "subsystem": subsystem,
                "moduleName": module_name,
                "description": desc,
            }
        )
    return rows


def filter_collect(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [r for r in rows if "数据资源采集汇聚" in r.get("subsystem", "")]


def filter_application(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        r for r in rows
        if r.get("system") == "数据共享交换平台-应用平台"
        and 44 <= int(r.get("rowIndex", 0)) <= 62
    ]


def dedupe_modules(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Merge continuation rows (empty module change, extra description) by moduleName."""
    out: list[dict[str, Any]] = []
    by_name: dict[str, dict[str, Any]] = {}
    for r in rows:
        name = r["moduleName"]
        if name not in by_name:
            by_name[name] = dict(r)
            out.append(by_name[name])
        else:
            prev = by_name[name]["description"]
            extra = r["description"]
            if extra and extra not in prev:
                by_name[name]["description"] = (prev + "\n" + extra).strip()
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--scope", default="all", choices=["all", "collect", "application"])
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    rows = load_rows()
    if args.scope == "collect":
        rows = filter_collect(rows)
        rows = dedupe_modules(rows)
        out_path = args.out or os.path.join(CATALOG_DIR, "v3-catalog.collect.json")
    elif args.scope == "application":
        rows = filter_application(rows)
        out_path = args.out or os.path.join(CATALOG_DIR, "v3-catalog.application.json")
    else:
        out_path = args.out or os.path.join(CATALOG_DIR, "v3-catalog.normalized.json")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    payload = {
        "source": "docs/requirements/功能清单.xlsx",
        "scope": args.scope,
        "rowCount": len(rows),
        "rows": rows,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
